"""
Exportador a formato Excel (.xlsx).
Genera archivos Excel con formato profesional.
"""

import logging
from pathlib import Path
from typing import List, Dict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exporta partidas a Excel con formato"""

    @staticmethod
    def exportar(partidas: List[Dict], output_path: str, nombre_hoja: str = 'Mediciones') -> None:
        """
        Exporta lista de partidas a Excel

        Args:
            partidas: lista de dicts con partidas
            output_path: ruta del archivo de salida
            nombre_hoja: nombre de la hoja
        """
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        try:
            # Crear DataFrame
            df = pd.DataFrame(partidas)

            # Renombrar columnas a nombres más descriptivos
            df = df.rename(columns={
                'capitulo': 'Capítulo',
                'subcapitulo': 'Subcapítulo',
                'apartado': 'Apartado',
                'codigo': 'Código Partida',
                'unidad': 'Unidad',
                'resumen': 'Título Partida',
                'descripcion': 'Descripción Partida',
                'cantidad': 'Unidades',
                'precio': 'Precio',
                'importe': 'Total'
            })

            # Reordenar columnas según especificación
            columnas_orden = [
                'Capítulo',
                'Subcapítulo',
                'Código Partida',
                'Unidad',
                'Título Partida',
                'Descripción Partida',
                'Unidades',
                'Precio',
                'Total'
            ]
            # Solo incluir columnas que existen
            columnas_existentes = [c for c in columnas_orden if c in df.columns]
            df = df[columnas_existentes]

            # Exportar a Excel
            df.to_excel(output_path, sheet_name=nombre_hoja, index=False, engine='openpyxl')

            # Aplicar formato
            ExcelExporter._aplicar_formato(output_path, nombre_hoja)

            logger.info(f"✓ Excel exportado: {output_path} ({len(partidas)} partidas)")

        except Exception as e:
            logger.error(f"Error exportando Excel: {e}")
            raise

    @staticmethod
    def _aplicar_formato(file_path: str, sheet_name: str) -> None:
        """Aplica formato profesional al Excel"""
        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name]

            # Estilos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Aplicar estilo a encabezados
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Ajustar ancho de columnas (según orden sin Apartado)
            column_widths = {
                'A': 12,  # Capítulo
                'B': 15,  # Subcapítulo
                'C': 15,  # Código Partida
                'D': 10,  # Unidad
                'E': 50,  # Título Partida
                'F': 60,  # Descripción Partida
                'G': 12,  # Unidades
                'H': 12,  # Precio
                'I': 15   # Total
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # Aplicar bordes y alineación a datos
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = border
                    if cell.column_letter in ['G', 'H', 'I']:  # Números (Unidades, Precio, Total)
                        cell.alignment = Alignment(horizontal='right')
                        if cell.value:
                            cell.number_format = '#,##0.00'

            # Filtros
            ws.auto_filter.ref = ws.dimensions

            wb.save(file_path)

        except Exception as e:
            logger.warning(f"No se pudo aplicar formato: {e}")

    @staticmethod
    def exportar_multihojas(estructura: Dict, output_path: str) -> None:
        """
        Exporta estructura completa en múltiples hojas

        Args:
            estructura: dict con estructura completa
            output_path: ruta del archivo de salida
        """
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Hoja resumen
                resumen_data = []
                for cap in estructura.get('capitulos', []):
                    resumen_data.append({
                        'Nivel': 'CAPÍTULO',
                        'Código': cap['codigo'],
                        'Nombre': cap['nombre'],
                        'Subcapítulos': len(cap['subcapitulos'])
                    })
                    for sub in cap['subcapitulos']:
                        resumen_data.append({
                            'Nivel': 'SUBCAPÍTULO',
                            'Código': sub['codigo'],
                            'Nombre': sub['nombre'],
                            'Partidas': len(sub['partidas']) + sum(len(a['partidas']) for a in sub.get('apartados', []))
                        })

                df_resumen = pd.DataFrame(resumen_data)
                df_resumen.to_excel(writer, sheet_name='Resumen', index=False)

                # Hoja de todas las partidas
                todas_partidas = []
                for cap in estructura.get('capitulos', []):
                    for sub in cap.get('subcapitulos', []):
                        for partida in sub.get('partidas', []):
                            todas_partidas.append({
                                **partida,
                                'capitulo': cap['codigo'],
                                'subcapitulo': sub['codigo'],
                                'apartado': None
                            })
                        for apt in sub.get('apartados', []):
                            for partida in apt.get('partidas', []):
                                todas_partidas.append({
                                    **partida,
                                    'capitulo': cap['codigo'],
                                    'subcapitulo': sub['codigo'],
                                    'apartado': apt['codigo']
                                })

                # Crear DataFrame y renombrar columnas
                df_partidas = pd.DataFrame(todas_partidas)

                df_partidas = df_partidas.rename(columns={
                    'capitulo': 'Capítulo',
                    'subcapitulo': 'Subcapítulo',
                    'apartado': 'Apartado',
                    'codigo': 'Código Partida',
                    'unidad': 'Unidad',
                    'resumen': 'Título Partida',
                    'descripcion': 'Descripción Partida',
                    'cantidad': 'Unidades',
                    'precio': 'Precio',
                    'importe': 'Total'
                })

                # Reordenar columnas
                columnas_orden = [
                    'Capítulo',
                    'Subcapítulo',
                    'Código Partida',
                    'Unidad',
                    'Título Partida',
                    'Descripción Partida',
                    'Unidades',
                    'Precio',
                    'Total'
                ]
                columnas_existentes = [c for c in columnas_orden if c in df_partidas.columns]
                df_partidas = df_partidas[columnas_existentes]

                df_partidas.to_excel(writer, sheet_name='Partidas', index=False)

            # Aplicar formato a ambas hojas
            ExcelExporter._aplicar_formato(output_path, 'Resumen')
            ExcelExporter._aplicar_formato(output_path, 'Partidas')

            logger.info(f"✓ Excel multihojas exportado: {output_path}")

        except Exception as e:
            logger.error(f"Error exportando Excel multihojas: {e}")
            raise


if __name__ == "__main__":
    # Test
    partidas_test = [
        {
            'codigo': 'DEM06',
            'unidad': 'm',
            'resumen': 'CORTE PAVIMENTO EXISTENTE',
            'descripcion': 'Corte de pavimento...',
            'cantidad': 630.0,
            'precio': 1.12,
            'importe': 705.60
        }
    ]

    ExcelExporter.exportar(partidas_test, 'data/test_export.xlsx')
    print("✓ Test Excel completado")
